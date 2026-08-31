#!/usr/bin/env python3
"""Pre vs post weekend-MSBD enablement: speed, IFR, and delivery reliability."""

from __future__ import annotations

import os
import sys
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

_default_creds = ROOT / ".gcp" / "credentials.json"
if _default_creds.exists():
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = str(_default_creds)

from gbq import query_df

SQL_DIR = ROOT / "sql"
OUT = ROOT / "output" / "weekend_shipping_pre_post"
CHARTS = ROOT / "docs" / "small_parcel" / "weekend_shipping_pre_post_charts"
DOC = ROOT / "docs" / "small_parcel" / "weekend_shipping_pre_post_impact.md"

NAVY = "#1a365d"
ACCENT = "#2e86ab"
ORANGE = "#e67e22"
RED = "#c0392b"
GREEN = "#1e8449"
GRAY = "#7f8c8d"
GOLD = "#d4a017"

POST_CUTOFF = pd.Timestamp("2026-08-17")


def load_sql(name: str) -> str:
    return (SQL_DIR / name).read_text()


def with_cte(select_file: str) -> str:
    return load_sql("weekend_shipping_pre_post_cte.sql") + "\n" + load_sql(select_file)


def prep_orders(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["order_complete_date"] = pd.to_datetime(df["order_complete_date"])
    df["msbd_su"] = pd.to_datetime(df["msbd_su"])
    df["enable_week"] = pd.to_datetime(df["enable_week"])
    df["week_start"] = pd.to_datetime(df["week_start"])
    df["delivery_date"] = pd.to_datetime(df["delivery_date"])
    # pandas dayofweek: Mon=0 … Sun=6. Weekend MSBD = Sat/Sun = 5, 6.
    df["is_weekend_msbd"] = df["msbd_su"].dt.dayofweek.isin([5, 6]).astype(float)
    df["is_weekend_ship"] = df["induction_dow_adj"].isin([1, 7]).astype(float)
    df["is_sat_ship"] = (df["induction_dow_adj"] == 7).astype(float)
    df["is_sun_ship"] = (df["induction_dow_adj"] == 1).astype(float)
    df["ifr"] = pd.to_numeric(df["inducted_on_time_or_early"], errors="coerce")
    df["o2d_stated"] = pd.to_numeric(df["o2d_stated"], errors="coerce")
    df["o2d_actual"] = pd.to_numeric(df["o2d_actual"], errors="coerce")
    df["o2s_actual"] = pd.to_numeric(df["o2s_actual"], errors="coerce")
    df["o2s_stated"] = pd.to_numeric(df["o2sumsbd"], errors="coerce")
    df["o2s_stated_1"] = pd.to_numeric(df["o2s_stated_1"], errors="coerce")
    df["fast_badge"] = pd.to_numeric(df["o2d_stated_5"], errors="coerce")
    # Delivery outcome, not a badge — Fast badge is stated O2D only.
    df["o2d_actual_less5"] = pd.to_numeric(df["o2d_actual_5"], errors="coerce")
    df["delivery_rel"] = pd.to_numeric(df["delivery_rel"], errors="coerce")
    df["det_delivery_date"] = pd.to_datetime(df["det_delivery_date"], errors="coerce")
    df["o2d_det"] = (df["det_delivery_date"] - df["order_complete_date"]).dt.days
    df["det_del_rel"] = pd.to_numeric(df["det_del_rel"], errors="coerce")
    return df[df["period"].notna() & df["order_bucket"].notna()].copy()


def summarize(g: pd.DataFrame) -> pd.Series:
    delivered = g["delivery_date"].notna()
    has_det = g["det_delivery_date"].notna() if "det_delivery_date" in g.columns else pd.Series(False, index=g.index)
    vol = int(g["ops"].nunique()) if "ops" in g.columns else int(len(g))
    n_suppliers = int(g["supplier_id"].nunique()) if "supplier_id" in g.columns else 1
    ifr = float(g["ifr"].mean()) if vol else np.nan
    return pd.Series(
        {
            "n_suppliers": n_suppliers,
            "vol": vol,
            "weekend_msbd": float(g["is_weekend_msbd"].mean()) if vol else np.nan,
            "weekend_ship": float(g["is_weekend_ship"].mean()) if vol else np.nan,
            "sat_ship": float(g["is_sat_ship"].mean()) if vol else np.nan,
            "sun_ship": float(g["is_sun_ship"].mean()) if vol else np.nan,
            "ifr": ifr,
            "late_orders": vol * (1 - ifr) if vol and pd.notna(ifr) else np.nan,
            "o2s_stated": float(g["o2s_stated"].mean()) if vol else np.nan,
            "o2s_stated_1": float(g["o2s_stated_1"].mean()) if vol else np.nan,
            "o2d_stated": float(g["o2d_stated"].mean()) if vol else np.nan,
            "o2d_actual": float(g["o2d_actual"].mean()) if vol else np.nan,
            "o2s_actual": float(g["o2s_actual"].mean()) if vol else np.nan,
            "fast_badge": float(g["fast_badge"].mean()) if vol else np.nan,
            "o2d_actual_less5": float(g["o2d_actual_less5"].mean()) if vol else np.nan,
            "o2d_det": float(g.loc[has_det, "o2d_det"].mean()) if has_det.any() else np.nan,
            "det_del_rel": float(g.loc[has_det, "det_del_rel"].mean()) if has_det.any() else np.nan,
            "del_rel": float(g.loc[delivered, "delivery_rel"].mean()) if delivered.any() else np.nan,
            "delivered_vol": int(g.loc[delivered, "ops"].nunique()),
            "det_vol": int(g.loc[has_det, "ops"].nunique()) if "ops" in g.columns else int(has_det.sum()),
        }
    )


def pct(x: float | None, digits: int = 1) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return "—"
    return f"{x * 100:.{digits}f}%"


def num(x: float | None, digits: int = 2) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return "—"
    return f"{x:.{digits}f}"


def pp(post: float, pre: float, digits: int = 1) -> str:
    if any(pd.isna(v) for v in (post, pre)):
        return "—"
    d = (post - pre) * 100
    sign = "+" if d >= 0 else ""
    return f"{sign}{d:.{digits}f} pp"


def days(post: float, pre: float, digits: int = 2) -> str:
    if any(pd.isna(v) for v in (post, pre)):
        return "—"
    d = post - pre
    sign = "+" if d >= 0 else ""
    return f"{sign}{d:.{digits}f}d"


def slice_pair(rollup: pd.DataFrame, wave: str, bucket: str) -> tuple[pd.Series, pd.Series]:
    sub = rollup[(rollup["wave"] == wave) & (rollup["order_bucket"] == bucket)]
    pre = sub[sub["period"] == "pre"].iloc[0]
    post = sub[sub["period"] == "post"].iloc[0]
    return pre, post


def metric_table_md(pre: pd.Series, post: pd.Series) -> str:
    rows = [
        ("Volume (distinct ops)", f"{int(pre['vol']):,}", f"{int(post['vol']):,}", "—"),
        ("Weekend MSBD share", pct(pre["weekend_msbd"]), pct(post["weekend_msbd"]), pp(post["weekend_msbd"], pre["weekend_msbd"])),
        ("Weekend ship (Sat+Sun adj)", pct(pre["weekend_ship"]), pct(post["weekend_ship"]), pp(post["weekend_ship"], pre["weekend_ship"])),
        ("Saturday induction", pct(pre["sat_ship"]), pct(post["sat_ship"]), pp(post["sat_ship"], pre["sat_ship"])),
        ("Sunday induction", pct(pre["sun_ship"]), pct(post["sun_ship"]), pp(post["sun_ship"], pre["sun_ship"])),
        ("IFR (on-time vs supplier MSBD)", pct(pre["ifr"]), pct(post["ifr"]), pp(post["ifr"], pre["ifr"])),
        ("Late orders", f"{pre['late_orders']:,.0f}", f"{post['late_orders']:,.0f}", f"{post['late_orders'] - pre['late_orders']:+,.0f}"),
        ("Stated O2S (days, o2sumsbd)", num(pre["o2s_stated"]), num(post["o2s_stated"]), days(post["o2s_stated"], pre["o2s_stated"])),
        ("1-day stated O2S share", pct(pre["o2s_stated_1"]), pct(post["o2s_stated_1"]), pp(post["o2s_stated_1"], pre["o2s_stated_1"])),
        ("Stated O2D (days)", num(pre["o2d_stated"]), num(post["o2d_stated"]), days(post["o2d_stated"], pre["o2d_stated"])),
        ("Actual O2D (days)", num(pre["o2d_actual"]), num(post["o2d_actual"]), days(post["o2d_actual"], pre["o2d_actual"])),
        ("Actual O2S (days)", num(pre["o2s_actual"]), num(post["o2s_actual"]), days(post["o2s_actual"], pre["o2s_actual"])),
        ("Fast badge (stated O2D ≤ 5)", pct(pre["fast_badge"]), pct(post["fast_badge"]), pp(post["fast_badge"], pre["fast_badge"])),
        ("Actual O2D ≤ 5", pct(pre["o2d_actual_less5"]), pct(post["o2d_actual_less5"]), pp(post["o2d_actual_less5"], pre["o2d_actual_less5"])),
        ("Deterministic O2D (days)", num(pre["o2d_det"]), num(post["o2d_det"]), days(post["o2d_det"], pre["o2d_det"])),
        ("Deterministic reliability", pct(pre["det_del_rel"]), pct(post["det_del_rel"]), pp(post["det_del_rel"], pre["det_del_rel"])),
        ("Delivery reliability", pct(pre["del_rel"]), pct(post["del_rel"]), pp(post["del_rel"], pre["del_rel"])),
    ]
    lines = [
        "| Metric | Pre | Post | Change |",
        "| --- | ---: | ---: | ---: |",
    ]
    for label, a, b, c in rows:
        lines.append(f"| {label} | {a} | {b} | {c} |")
    return "\n".join(lines)


def chart_grouped_bars(wave1: pd.DataFrame) -> None:
    pre, post = slice_pair(wave1, "wave1_jun_jul", "fri_sat")
    labels = ["Weekend\nMSBD", "Weekend\nship", "IFR", "Fast\nbadge", "Delivery\nrel"]
    keys = ["weekend_msbd", "weekend_ship", "ifr", "fast_badge", "del_rel"]
    pre_v = [pre[k] * 100 for k in keys]
    post_v = [post[k] * 100 for k in keys]

    fig, ax = plt.subplots(figsize=(10, 5.5))
    x = np.arange(len(labels))
    w = 0.36
    ax.bar(x - w / 2, pre_v, w, label="Pre (6w before enable)", color=GRAY)
    ax.bar(x + w / 2, post_v, w, label="Post (enable week through 8/16)", color=ACCENT)
    for i, (a, b) in enumerate(zip(pre_v, post_v)):
        ax.text(i - w / 2, a + 1.2, f"{a:.1f}%", ha="center", va="bottom", fontsize=8)
        ax.text(i + w / 2, b + 1.2, f"{b:.1f}%", ha="center", va="bottom", fontsize=8)
    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.set_ylim(0, 110)
    ax.set_ylabel("Percent")
    ax.set_title(
        "Wave 1 (12 suppliers, June 21 / July 5 enablement)\n"
        "Friday–Saturday placed orders — reliability vs promised speed mix"
    )
    ax.legend(frameon=False)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    fig.savefig(CHARTS / "01_wave1_fri_sat_rates.png", dpi=150)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(8, 5))
    speed_labels = ["Stated O2D", "Actual O2D", "Actual O2S"]
    speed_keys = ["o2d_stated", "o2d_actual", "o2s_actual"]
    pre_s = [pre[k] for k in speed_keys]
    post_s = [post[k] for k in speed_keys]
    x = np.arange(len(speed_labels))
    ax.bar(x - w / 2, pre_s, w, label="Pre", color=GRAY)
    ax.bar(x + w / 2, post_s, w, label="Post", color=GREEN)
    for i, (a, b) in enumerate(zip(pre_s, post_s)):
        ax.text(i - w / 2, a + 0.05, f"{a:.2f}", ha="center", fontsize=8)
        ax.text(i + w / 2, b + 0.05, f"{b:.2f}", ha="center", fontsize=8)
        ax.annotate(
            f"{b - a:+.2f}d",
            xy=(i + w / 2, b),
            xytext=(18, 8),
            textcoords="offset points",
            fontsize=8,
            color=GREEN if b < a else RED,
        )
    ax.set_xticks(x)
    ax.set_xticklabels(speed_labels)
    ax.set_ylabel("Days")
    ax.set_title("Wave 1 Fri/Sat orders — speed gain (lower is faster)")
    ax.legend(frameon=False)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    fig.savefig(CHARTS / "02_wave1_fri_sat_speed.png", dpi=150)
    plt.close(fig)


def chart_weekday_vs_weekend(rollup: pd.DataFrame) -> None:
    fig, axes = plt.subplots(1, 2, figsize=(11, 5))
    for ax, metric, title in (
        (axes[0], "ifr", "IFR"),
        (axes[1], "del_rel", "Delivery reliability"),
    ):
        x = np.arange(2)
        w = 0.35
        for i, bucket, color in ((0, "fri_sat", ACCENT), (1, "weekday", NAVY)):
            pre, post = slice_pair(rollup, "wave1_jun_jul", bucket)
            vals = [pre[metric] * 100, post[metric] * 100]
            offset = -w / 2 if bucket == "fri_sat" else w / 2
            label = "Fri/Sat placed" if bucket == "fri_sat" else "Mon–Thu placed"
            ax.plot(
                x,
                vals,
                marker="o",
                linewidth=2.4,
                color=color,
                label=label,
            )
            for xi, v in zip(x, vals):
                ax.text(xi, v + 0.8, f"{v:.1f}%", ha="center", fontsize=8, color=color)
        ax.set_xticks(x)
        ax.set_xticklabels(["Pre", "Post"])
        ax.set_title(title)
        ax.set_ylim(70, 100)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.legend(frameon=False, loc="lower left")
    fig.suptitle(
        "Wave 1 — reliability hit is isolated to weekend-placed orders\n"
        "Same 12 enabled warehouses; weekday IFR/del-rel held or improved",
        y=1.03,
    )
    fig.tight_layout()
    fig.savefig(CHARTS / "03_wave1_weekday_vs_weekend_reliability.png", dpi=150, bbox_inches="tight")
    plt.close(fig)


def chart_sat_sun(rollup: pd.DataFrame) -> None:
    pre, post = slice_pair(rollup, "wave1_jun_jul", "fri_sat")
    fig, ax = plt.subplots(figsize=(8, 5))
    x = np.arange(2)
    w = 0.35
    sat = [pre["sat_ship"] * 100, post["sat_ship"] * 100]
    sun = [pre["sun_ship"] * 100, post["sun_ship"] * 100]
    ax.bar(x - w / 2, sat, w, label="Saturday induction", color=ORANGE)
    ax.bar(x + w / 2, sun, w, label="Sunday induction", color=NAVY)
    for i, (a, b) in enumerate(zip(sat, sun)):
        ax.text(i - w / 2, a + 1, f"{a:.1f}%", ha="center", fontsize=8)
        ax.text(i + w / 2, b + 1, f"{b:.1f}%", ha="center", fontsize=8)
    ax.set_xticks(x)
    ax.set_xticklabels(["Pre", "Post"])
    ax.set_ylabel("% of Fri/Sat placed orders")
    ax.set_ylim(0, 85)
    ax.set_title(
        "Wave 1 Fri/Sat orders — still almost no Saturday induction\n"
        "Sunday ship held (~65%); Saturday MSBD is what IFR is missing"
    )
    ax.legend(frameon=False)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    fig.savefig(CHARTS / "04_wave1_sat_vs_sun_induction.png", dpi=150)
    plt.close(fig)


def chart_supplier_lollipop(supplier: pd.DataFrame) -> None:
    sub = supplier[
        (supplier["wave"] == "wave1_jun_jul")
        & (supplier["order_bucket"] == "fri_sat")
    ].copy()
    pre = sub[sub["period"] == "pre"].set_index("supplier_id")
    post = sub[sub["period"] == "post"].set_index("supplier_id")
    names = post["su_name"].to_dict()
    rows = []
    for sid in post.index:
        if sid not in pre.index:
            continue
        rows.append(
            {
                "name": names[sid],
                "ifr_delta": (post.loc[sid, "ifr"] - pre.loc[sid, "ifr"]) * 100,
                "o2d_act_delta": post.loc[sid, "o2d_actual"] - pre.loc[sid, "o2d_actual"],
                "del_delta": (post.loc[sid, "del_rel"] - pre.loc[sid, "del_rel"]) * 100,
                "post_vol": post.loc[sid, "vol"],
            }
        )
    plot = pd.DataFrame(rows).sort_values("ifr_delta")
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = [RED if v < 0 else GREEN for v in plot["ifr_delta"]]
    ax.hlines(plot["name"], 0, plot["ifr_delta"], color=colors, linewidth=2)
    ax.scatter(plot["ifr_delta"], plot["name"], color=colors, zorder=3)
    ax.axvline(0, color="#333", linewidth=0.8)
    ax.set_xlabel("IFR change (pp), Fri/Sat placed")
    ax.set_title("Wave 1 warehouses — IFR change on weekend-placed orders")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    fig.savefig(CHARTS / "05_wave1_supplier_ifr_change.png", dpi=150)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(10, 6))
    colors = [GREEN if v < 0 else RED for v in plot["o2d_act_delta"]]
    plot2 = plot.sort_values("o2d_act_delta")
    ax.hlines(plot2["name"], 0, plot2["o2d_act_delta"], color=colors, linewidth=2)
    ax.scatter(plot2["o2d_act_delta"], plot2["name"], color=colors, zorder=3)
    ax.axvline(0, color="#333", linewidth=0.8)
    ax.set_xlabel("Actual O2D change (days); negative = faster")
    ax.set_title("Wave 1 warehouses — actual speed change on weekend-placed orders")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    fig.savefig(CHARTS / "06_wave1_supplier_o2d_actual_change.png", dpi=150)
    plt.close(fig)


def chart_weekly(orders: pd.DataFrame) -> None:
    w1 = orders[(orders["wave"] == "wave1_jun_jul") & (orders["order_bucket"] == "fri_sat")]
    weekly = (
        w1.groupby("weeks_from_enable", dropna=True)
        .apply(summarize)
        .reset_index()
    )
    weekly = weekly[weekly["vol"] >= 50]
    fig, axes = plt.subplots(2, 1, figsize=(11, 8), sharex=True)
    axes[0].plot(weekly["weeks_from_enable"], weekly["ifr"] * 100, marker="o", color=RED, label="IFR")
    axes[0].plot(
        weekly["weeks_from_enable"], weekly["del_rel"] * 100, marker="s", color=NAVY, label="Delivery rel"
    )
    axes[0].axvline(0, color=ORANGE, linestyle="--", label="Enable week")
    axes[0].set_ylabel("Percent")
    axes[0].set_ylim(50, 100)
    axes[0].legend(frameon=False)
    axes[0].set_title("Wave 1 Fri/Sat — IFR and delivery reliability by weeks from enablement")
    axes[1].plot(
        weekly["weeks_from_enable"], weekly["o2d_stated"], marker="o", color=ACCENT, label="Stated O2D"
    )
    axes[1].plot(
        weekly["weeks_from_enable"], weekly["o2d_actual"], marker="s", color=GREEN, label="Actual O2D"
    )
    axes[1].axvline(0, color=ORANGE, linestyle="--")
    axes[1].set_ylabel("Days")
    axes[1].set_xlabel("Weeks from supplier enablement week (0 = first weekend-MSBD week)")
    axes[1].legend(frameon=False)
    axes[1].set_title("Wave 1 Fri/Sat — stated vs actual O2D")
    for ax in axes:
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
    fig.tight_layout()
    fig.savefig(CHARTS / "07_wave1_weekly_relative.png", dpi=150)
    plt.close(fig)


def chart_wave2(rollup: pd.DataFrame) -> None:
    fig, axes = plt.subplots(1, 2, figsize=(11, 5))
    w = 0.35
    for ax, bucket, title in (
        (axes[0], "fri_sat", "Fri/Sat placed"),
        (axes[1], "weekday", "Mon–Thu placed"),
    ):
        pre, post = slice_pair(rollup, "wave2_aug", bucket)
        labels = ["IFR", "Del rel", "Fast badge"]
        keys = ["ifr", "del_rel", "fast_badge"]
        x = np.arange(len(labels))
        ax.bar(x - w / 2, [pre[k] * 100 for k in keys], w, color=GRAY, label="Pre")
        ax.bar(x + w / 2, [post[k] * 100 for k in keys], w, color=GOLD, label="Post (early)")
        ax.set_xticks(x)
        ax.set_xticklabels(labels)
        ax.set_ylim(0, 110)
        ax.set_title(title)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.legend(frameon=False)
    fig.suptitle("Wave 2 (August 2 Q3 sprint, ~70 suppliers) — early 2-week post window")
    fig.tight_layout()
    fig.savefig(CHARTS / "08_wave2_early_read.png", dpi=150)
    plt.close(fig)


def supplier_md(supplier: pd.DataFrame, wave: str) -> str:
    sub = supplier[(supplier["wave"] == wave) & (supplier["order_bucket"] == "fri_sat")].copy()
    pre = sub[sub["period"] == "pre"].set_index("supplier_id")
    post = sub[sub["period"] == "post"].set_index("supplier_id")
    lines = [
        "| Warehouse | Enable week | Pre vol | Post vol | IFR pre → post | Del rel pre → post | Stated O2D | Actual O2D | Sat ship post | Sun ship post |",
        "| --- | --- | ---: | ---: | --- | --- | --- | --- | ---: | ---: |",
    ]
    for sid in post.sort_values("vol", ascending=False).index:
        if sid not in pre.index:
            continue
        p, t = pre.loc[sid], post.loc[sid]
        enable = pd.to_datetime(t["enable_week"]).date()
        lines.append(
            "| {name} | {en} | {pv:,} | {tv:,} | {ip} → {it} ({idlt}) | {dp} → {dt} ({dd}) | "
            "{sp} → {st} ({sd}) | {ap} → {at} ({ad}) | {sat} | {sun} |".format(
                name=t["su_name"],
                en=enable,
                pv=int(p["vol"]),
                tv=int(t["vol"]),
                ip=pct(p["ifr"]),
                it=pct(t["ifr"]),
                idlt=pp(t["ifr"], p["ifr"]),
                dp=pct(p["del_rel"]),
                dt=pct(t["del_rel"]),
                dd=pp(t["del_rel"], p["del_rel"]),
                sp=num(p["o2d_stated"]),
                st=num(t["o2d_stated"]),
                sd=days(t["o2d_stated"], p["o2d_stated"]),
                ap=num(p["o2d_actual"]),
                at=num(t["o2d_actual"]),
                ad=days(t["o2d_actual"], p["o2d_actual"]),
                sat=pct(t["sat_ship"]),
                sun=pct(t["sun_ship"]),
            )
        )
    return "\n".join(lines)


def write_report(
    roster: pd.DataFrame,
    rollup: pd.DataFrame,
    supplier: pd.DataFrame,
    control: pd.DataFrame,
) -> None:
    w1_fs_pre, w1_fs_post = slice_pair(rollup, "wave1_jun_jul", "fri_sat")
    w1_wd_pre, w1_wd_post = slice_pair(rollup, "wave1_jun_jul", "weekday")
    w1_all_pre, w1_all_post = slice_pair(rollup, "wave1_jun_jul", "all")
    w2_fs_pre, w2_fs_post = slice_pair(rollup, "wave2_aug", "fri_sat")
    w2_all_pre, w2_all_post = slice_pair(rollup, "wave2_aug", "all")

    ctrl = control.copy()
    ctrl_fs = ctrl[ctrl["order_bucket"] == "fri_sat"].set_index("period")

    n_w1 = int(roster[roster["wave"] == "wave1_jun_jul"]["supplier_id"].nunique())
    n_w2 = int(roster[roster["wave"] == "wave2_aug"]["supplier_id"].nunique())

    md = f"""# Weekend shipping enablement — pre vs post impact

Generated: 2026-08-31

June–July Wave 1 enabled **weekend Must Ship By Date** for 24-hour dropship warehouses that were already inducting Fri/Sat orders over the weekend. This rerun measures whether those warehouses are performing, the speed gain, and delivery reliability on **Friday–Saturday placed orders**.

An August 2 Q3 sprint wave is included as an early read (short post window).

## Bottom line

Wave 1 (12 warehouses; nuLOOM NJ on **2026-06-21**, then Safavieh / Aosom / GigaCloud / SLM / WINADO / nuLOOM CA on **2026-07-05 / 07-12**):

- **Setting took:** Fri/Sat weekend MSBD share went from {pct(w1_fs_pre['weekend_msbd'])} to {pct(w1_fs_post['weekend_msbd'])}.
- **Speed improved, but less than the promise:** stated O2D {num(w1_fs_pre['o2d_stated'])} → {num(w1_fs_post['o2d_stated'])} days ({days(w1_fs_post['o2d_stated'], w1_fs_pre['o2d_stated'])}); actual O2D {num(w1_fs_pre['o2d_actual'])} → {num(w1_fs_post['o2d_actual'])} ({days(w1_fs_post['o2d_actual'], w1_fs_pre['o2d_actual'])}). Fast badge {pct(w1_fs_pre['fast_badge'])} → {pct(w1_fs_post['fast_badge'])} ({pp(w1_fs_post['fast_badge'], w1_fs_pre['fast_badge'])}).
- **Excel:** all warehouse-level pre/post rows plus this summary are in `output/weekend_shipping_pre_post/weekend_shipping_pre_post_supplier.xlsx`.
- **Reliability got worse on weekend-placed orders:** IFR {pct(w1_fs_pre['ifr'])} → {pct(w1_fs_post['ifr'])} ({pp(w1_fs_post['ifr'], w1_fs_pre['ifr'])}); delivery reliability {pct(w1_fs_pre['del_rel'])} → {pct(w1_fs_post['del_rel'])} ({pp(w1_fs_post['del_rel'], w1_fs_pre['del_rel'])}).
- **Operations did not move to Saturday:** weekend induction held ({pct(w1_fs_pre['weekend_ship'])} → {pct(w1_fs_post['weekend_ship'])}), but it is almost all **Sunday** ({pct(w1_fs_post['sun_ship'])}) vs **Saturday** ({pct(w1_fs_post['sat_ship'])}). Friday orders with a Saturday MSBD that induct Sunday miss IFR.
- **Weekdays for the same warehouses improved:** IFR {pct(w1_wd_pre['ifr'])} → {pct(w1_wd_post['ifr'])}; delivery reliability {pct(w1_wd_pre['del_rel'])} → {pct(w1_wd_post['del_rel'])}. The miss is weekend-order specific.

Wave 2 ({n_w2} warehouses, enable week **2026-08-02**; post is only ~2 weeks through 8/16) shows the same pattern: promised speed and fast-badge up, IFR and delivery reliability down on Fri/Sat orders.

## Wave 1 — overall stated speed and badging (all order days)

This is the warehouse-level mix including weekdays, so the stated-speed move is smaller than the Fri/Sat-only slice, but it is the right view for “did promises get faster overall.”

{metric_table_md(w1_all_pre, w1_all_post)}

## Wave 1 — Friday/Saturday placed orders

{n_w1} 24-hour DS warehouses. Pre = 6 weeks before each warehouse’s first weekend-MSBD week. Post = enable week through 2026-08-16 (14-day delivery lag).

{metric_table_md(w1_fs_pre, w1_fs_post)}

### Same warehouses, Monday–Thursday placed (within-supplier control)

{metric_table_md(w1_wd_pre, w1_wd_post)}

## Wave 1 warehouse detail (Fri/Sat)

{supplier_md(supplier, "wave1_jun_jul")}

WINADO’s post weekend-MSBD share did not stay high after the first week — treat as not sustained.

## Wave 2 early read (August 2 Q3 sprint)

Post window is 2026-08-02 through 2026-08-16 only. Directionally the same as Wave 1 on Fri/Sat orders: weekend MSBD {pct(w2_fs_pre['weekend_msbd'])} → {pct(w2_fs_post['weekend_msbd'])}; IFR {pct(w2_fs_pre['ifr'])} → {pct(w2_fs_post['ifr'])}; delivery reliability {pct(w2_fs_pre['del_rel'])} → {pct(w2_fs_post['del_rel'])}; stated O2D {num(w2_fs_pre['o2d_stated'])} → {num(w2_fs_post['o2d_stated'])}; actual O2D {num(w2_fs_pre['o2d_actual'])} → {num(w2_fs_post['o2d_actual'])}. Saturday induction is higher than Wave 1 ({pct(w2_fs_post['sat_ship'])}) but still minority vs Sunday ({pct(w2_fs_post['sun_ship'])}).

### Wave 2 overall (all order days) — stated speed and badging

{metric_table_md(w2_all_pre, w2_all_post)}

{metric_table_md(w2_fs_pre, w2_fs_post)}

Full Wave 2 warehouse file: `output/weekend_shipping_pre_post/weekend_shipping_pre_post_supplier.xlsx` (sheets `Supplier_FriSat` and `Supplier_Overall`).

## 24hr control (not enabled)

24-hour DS suppliers with **no** Fri/Sat weekend MSBD on/after 2026-06-21. Calendar windows: pre 5/03–6/21, post 7/05–8/16.

"""

    if "pre" in ctrl_fs.index and "post" in ctrl_fs.index:
        cpre, cpost = ctrl_fs.loc["pre"], ctrl_fs.loc["post"]
        md += (
            f"Fri/Sat control: IFR {pct(cpre['ifr'])} → {pct(cpost['ifr'])} "
            f"({pp(cpost['ifr'], cpre['ifr'])}); delivery reliability "
            f"{pct(cpre['del_rel'])} → {pct(cpost['del_rel'])} "
            f"({pp(cpost['del_rel'], cpre['del_rel'])}); actual O2D "
            f"{num(cpre['o2d_actual'])} → {num(cpost['o2d_actual'])} "
            f"({days(cpost['o2d_actual'], cpre['o2d_actual'])}). "
            "Control reliability did not drop the way enabled Fri/Sat orders did.\n"
        )

    md += """
## What this means

1. **Enablement worked in the promise:** weekend MSBD is on; stated O2D and 5-day badging improved on Fri/Sat orders.
2. **Warehouses are not failing to ship weekends** — they still induct ~65% of Fri/Sat orders on Sunday. They are failing **Saturday** induction, which is what the new MSBD requires for Friday orders.
3. **Do not read IFR as “they stopped trying.”** Tighter Saturday MSBD converted prior Sunday-early (vs Monday MSBD) into Sunday-late (vs Saturday MSBD).
4. **Delivery reliability followed IFR down** on weekend-placed orders only. Weekday del-rel for the same buildings is stable/up.
5. **Next lever:** Saturday FedEx pickup / Saturday induction at Wave 1 buildings (especially Safavieh IN/TX/CA and NFusion El Monte) before adding more weekend-MSBD enablement. nuLOOM CA is the closest to a Saturday-ship success case.

## Methodology

| Item | Rule |
| --- | --- |
| Table | `` `wf-gcp-us-ae-global-tnd-prod.speed_and_reliability.HVE_perf_Monitoring` `` |
| Fulfillment | `fulfillment_type = 'DS'`, `sp_lt = 24` |
| Weekend-placed cohort | Friday + Saturday orders: `order_dow IN (6, 7)` |
| Weekday control | Monday–Thursday: `order_dow IN (2, 3, 4, 5)` |
| Day-of-week | **Sunday = 1 … Saturday = 7** for `order_dow` and `induction_dow_adj` (verified vs calendar dates) |
| Weekend ship | `induction_dow_adj IN (1, 7)` — not `inducted_over_weekend` (that flag undercounts Sunday scans) |
| Weekend MSBD | `EXTRACT(DAYOFWEEK FROM msbd_su) IN (1, 7)` |
| Enabled | First week on/after 2026-06-21 with ≥20 Fri/Sat ops and ≥10% weekend MSBD; prior 6-week weekend-MSBD average < 5% |
| Pre / post | Supplier-specific: 6 weeks before enable week vs enable week through 2026-08-16 |
| Volume | `COUNT(DISTINCT ops)` |
| IFR | `AVG(inducted_on_time_or_early)` |
| Delivery reliability | `AVG(delivery_rel)` among rows with `delivery_date IS NOT NULL` |
| Speed | Stated O2S = `AVG(o2sumsbd)`; stated O2D = `AVG(o2d_stated)`; actual O2S/O2D; 1-day O2S = `AVG(o2s_stated_1)`; Fast badge = `AVG(o2d_stated_5)` (stated only); Actual O2D ≤ 5 = `AVG(o2d_actual_5)` (outcome, not a badge) |

The July 2026 candidate finder (`sql/weekend_shipping_supplier_analysis.sql`) treated `order_dow IN (5, 6)` as Fri/Sat. Empirically that is **Thursday + Friday**. This impact rerun uses the corrected Friday + Saturday filter.

## Charts

![Wave 1 rates](weekend_shipping_pre_post_charts/01_wave1_fri_sat_rates.png)

![Wave 1 speed](weekend_shipping_pre_post_charts/02_wave1_fri_sat_speed.png)

![Weekday vs weekend reliability](weekend_shipping_pre_post_charts/03_wave1_weekday_vs_weekend_reliability.png)

![Sat vs Sun induction](weekend_shipping_pre_post_charts/04_wave1_sat_vs_sun_induction.png)

![Warehouse IFR change](weekend_shipping_pre_post_charts/05_wave1_supplier_ifr_change.png)

![Warehouse actual O2D change](weekend_shipping_pre_post_charts/06_wave1_supplier_o2d_actual_change.png)

![Weekly relative](weekend_shipping_pre_post_charts/07_wave1_weekly_relative.png)

![Wave 2 early](weekend_shipping_pre_post_charts/08_wave2_early_read.png)
"""
    DOC.write_text(md)


RATE_COLS = {
    "weekend_msbd",
    "weekend_ship",
    "sat_ship",
    "sun_ship",
    "ifr",
    "o2s_stated_1",
    "fast_badge",
    "o2d_actual_less5",
    "del_rel",
    "det_del_rel",
    "ifr_overall",
    "ifr_fri_sat",
    "ifr_weekday",
    "del_rel_overall",
    "del_rel_fri_sat",
    "del_rel_weekday",
    "det_del_rel_overall",
    "det_del_rel_fri_sat",
    "det_del_rel_weekday",
    "o2d_actual_less5_overall",
    "o2d_actual_less5_fri_sat",
    "o2d_actual_less5_weekday",
    "fri_sat_vol_share",
    "weekday_vol_share",
    "late_share_fri_sat",
    "weekend_ifr_hit_on_overall",
    "weekday_ifr_hit_on_overall",
    "weekend_del_rel_hit_on_overall",
    "weekday_del_rel_hit_on_overall",
    "overall_ifr_if_frisat_held",
    "overall_del_rel_if_frisat_held",
}
DAY_COLS = {
    "o2s_stated",
    "o2d_stated",
    "o2s_actual",
    "o2d_actual",
    "o2d_det",
    "o2d_det_overall",
    "o2d_det_fri_sat",
    "o2d_det_weekday",
}
COUNT_COLS = {
    "vol",
    "late_orders",
    "delivered_vol",
    "det_vol",
    "n_suppliers",
    "vol_overall",
    "vol_fri_sat",
    "vol_weekday",
    "late_orders_overall",
    "late_orders_fri_sat",
    "late_orders_weekday",
    "delivered_vol_overall",
    "delivered_vol_fri_sat",
    "delivered_vol_weekday",
    "extra_late_fri_sat",
    "extra_late_weekday",
}

WIDE_METRICS = [
    "vol",
    "o2s_stated",
    "o2s_stated_1",
    "o2d_stated",
    "fast_badge",
    "o2d_actual_less5",
    "o2s_actual",
    "o2d_actual",
    "o2d_det",
    "ifr",
    "del_rel",
    "det_del_rel",
    "late_orders",
    "delivered_vol",
    "weekend_msbd",
    "weekend_ship",
    "sat_ship",
    "sun_ship",
]
WIDE_ID = ["wave", "supplier_id", "su_name", "parent_su_name", "sto", "srm", "enable_week"]

# Fast badge is stated O2D only. o2d_actual_less5 is delivery outcome, not a badge.
EXCEL_HEADERS = {
    "fast_badge": "Fast badge (stated O2D ≤ 5)",
    "pre_fast_badge": "Pre Fast badge (stated O2D ≤ 5)",
    "post_fast_badge": "Post Fast badge (stated O2D ≤ 5)",
    "delta_fast_badge": "Δ Fast badge (stated O2D ≤ 5)",
    "pre_o2d_actual_less5": "pre_o2d_actual_less5",
    "post_o2d_actual_less5": "post_o2d_actual_less5",
    "delta_o2d_actual_less5": "delta_o2d_actual_less5",
    "o2d_det": "Deterministic O2D (days)",
    "pre_o2d_det": "Pre det O2D (days)",
    "post_o2d_det": "Post det O2D (days)",
    "delta_o2d_det": "Δ det O2D (days)",
    "det_del_rel": "Deterministic reliability",
    "pre_det_del_rel": "Pre det reliability",
    "post_det_del_rel": "Post det reliability",
    "delta_det_del_rel": "Δ det reliability",
    "pre_ifr_fri_sat": "Pre IFR Fri/Sat",
    "post_ifr_fri_sat": "Post IFR Fri/Sat",
    "delta_ifr_fri_sat": "Δ IFR Fri/Sat",
    "pre_ifr_weekday": "Pre IFR Mon–Thu",
    "post_ifr_weekday": "Post IFR Mon–Thu",
    "delta_ifr_weekday": "Δ IFR Mon–Thu",
    "pre_ifr_overall": "Pre IFR overall",
    "post_ifr_overall": "Post IFR overall",
    "delta_ifr_overall": "Δ IFR overall",
    "pre_del_rel_fri_sat": "Pre DR Fri/Sat",
    "post_del_rel_fri_sat": "Post DR Fri/Sat",
    "delta_del_rel_fri_sat": "Δ DR Fri/Sat",
    "pre_del_rel_weekday": "Pre DR Mon–Thu",
    "post_del_rel_weekday": "Post DR Mon–Thu",
    "delta_del_rel_weekday": "Δ DR Mon–Thu",
    "pre_del_rel_overall": "Pre DR overall",
    "post_del_rel_overall": "Post DR overall",
    "delta_del_rel_overall": "Δ DR overall",
    "pre_o2d_det_fri_sat": "Pre det O2D Fri/Sat",
    "post_o2d_det_fri_sat": "Post det O2D Fri/Sat",
    "delta_o2d_det_fri_sat": "Δ det O2D Fri/Sat",
    "pre_o2d_det_weekday": "Pre det O2D Mon–Thu",
    "post_o2d_det_weekday": "Post det O2D Mon–Thu",
    "delta_o2d_det_weekday": "Δ det O2D Mon–Thu",
    "pre_o2d_det_overall": "Pre det O2D overall",
    "post_o2d_det_overall": "Post det O2D overall",
    "delta_o2d_det_overall": "Δ det O2D overall",
    "pre_det_del_rel_fri_sat": "Pre det rel Fri/Sat",
    "post_det_del_rel_fri_sat": "Post det rel Fri/Sat",
    "delta_det_del_rel_fri_sat": "Δ det rel Fri/Sat",
    "pre_det_del_rel_weekday": "Pre det rel Mon–Thu",
    "post_det_del_rel_weekday": "Post det rel Mon–Thu",
    "delta_det_del_rel_weekday": "Δ det rel Mon–Thu",
    "pre_det_del_rel_overall": "Pre det rel overall",
    "post_det_del_rel_overall": "Post det rel overall",
    "delta_det_del_rel_overall": "Δ det rel overall",
    "post_fri_sat_vol_share": "Post Fri/Sat volume share",
    "post_weekday_vol_share": "Post Mon–Thu volume share",
    "post_late_share_fri_sat": "Post Fri/Sat share of lates",
    "overall_ifr_if_frisat_held": "Overall IFR if Fri/Sat IFR held",
    "weekend_ifr_hit_on_overall": "Fri/Sat IFR hit on overall",
    "weekday_ifr_hit_on_overall": "Mon–Thu IFR hit on overall",
    "overall_del_rel_if_frisat_held": "Overall DR if Fri/Sat DR held",
    "weekend_del_rel_hit_on_overall": "Fri/Sat DR hit on overall",
    "weekday_del_rel_hit_on_overall": "Mon–Thu DR hit on overall",
    "extra_late_fri_sat": "Extra Fri/Sat IFR lates vs pre",
    "extra_late_weekday": "Extra Mon–Thu IFR lates vs pre",
    "pre_vol_fri_sat": "Pre vol Fri/Sat",
    "post_vol_fri_sat": "Post vol Fri/Sat",
    "pre_vol_weekday": "Pre vol Mon–Thu",
    "post_vol_weekday": "Post vol Mon–Thu",
    "pre_vol_overall": "Pre vol overall",
    "post_vol_overall": "Post vol overall",
}


def wide_supplier(supplier: pd.DataFrame, bucket: str) -> pd.DataFrame:
    sub = supplier[supplier["order_bucket"] == bucket].copy()
    pre = sub[sub["period"] == "pre"]
    post = sub[sub["period"] == "post"]
    keep = [c for c in WIDE_ID + WIDE_METRICS if c in sub.columns]
    pre_r = pre[keep].rename(columns={m: f"pre_{m}" for m in WIDE_METRICS})
    post_r = post[["supplier_id"] + [m for m in WIDE_METRICS if m in post.columns]].rename(
        columns={m: f"post_{m}" for m in WIDE_METRICS}
    )
    out = pre_r.merge(post_r, on="supplier_id", how="outer")
    for m in WIDE_METRICS:
        pre_c, post_c = f"pre_{m}", f"post_{m}"
        if pre_c in out.columns and post_c in out.columns:
            out[f"delta_{m}"] = out[post_c] - out[pre_c]
    ordered = list(WIDE_ID)
    for m in WIDE_METRICS:
        ordered.extend([c for c in (f"pre_{m}", f"post_{m}", f"delta_{m}") if c in out.columns])
    out = out[[c for c in ordered if c in out.columns]]
    sort_col = "post_vol" if "post_vol" in out.columns else "pre_vol"
    out[sort_col] = pd.to_numeric(out[sort_col], errors="coerce")
    parts = []
    for wave in ("wave1_jun_jul", "wave2_aug"):
        parts.append(out[out["wave"] == wave].sort_values(sort_col, ascending=False, na_position="last"))
    parts.append(out[~out["wave"].isin(["wave1_jun_jul", "wave2_aug"])].sort_values(sort_col, ascending=False, na_position="last"))
    return pd.concat(parts, ignore_index=True)


IMPACT_METRICS = [
    "vol",
    "ifr",
    "del_rel",
    "o2d_det",
    "det_del_rel",
    "o2d_actual_less5",
    "late_orders",
    "delivered_vol",
]


def _wide_slice(wide: pd.DataFrame, suffix: str) -> pd.DataFrame:
    out = wide[WIDE_ID].copy()
    for m in IMPACT_METRICS:
        for p in ("pre", "post", "delta"):
            src = f"{p}_{m}"
            if src in wide.columns:
                out[f"{p}_{m}_{suffix}"] = wide[src]
    return out


def build_ifr_dr_impact(supplier: pd.DataFrame, supplier_all: pd.DataFrame) -> pd.DataFrame:
    """Mon–Thu vs Fri/Sat vs overall IFR/DR, plus weekend contribution to overall."""
    fs = _wide_slice(wide_supplier(supplier, "fri_sat"), "fri_sat")
    wd = _wide_slice(wide_supplier(supplier, "weekday"), "weekday")
    ov = _wide_slice(wide_supplier(supplier_all, "all"), "overall")
    out = ov.merge(fs.drop(columns=[c for c in WIDE_ID if c != "supplier_id"]), on="supplier_id", how="outer")
    out = out.merge(wd.drop(columns=[c for c in WIDE_ID if c != "supplier_id"]), on="supplier_id", how="outer")

    fs_vol = out["post_vol_fri_sat"]
    wd_vol = out["post_vol_weekday"]
    ov_vol = out["post_vol_overall"]
    out["post_fri_sat_vol_share"] = fs_vol / ov_vol
    out["post_weekday_vol_share"] = wd_vol / ov_vol
    out["weekend_ifr_hit_on_overall"] = out["post_fri_sat_vol_share"] * out["delta_ifr_fri_sat"]
    out["weekday_ifr_hit_on_overall"] = out["post_weekday_vol_share"] * out["delta_ifr_weekday"]
    out["overall_ifr_if_frisat_held"] = out["post_ifr_overall"] - out["weekend_ifr_hit_on_overall"]
    out["extra_late_fri_sat"] = fs_vol * (out["pre_ifr_fri_sat"] - out["post_ifr_fri_sat"])
    out["extra_late_weekday"] = wd_vol * (out["pre_ifr_weekday"] - out["post_ifr_weekday"])
    out["post_late_share_fri_sat"] = out["post_late_orders_fri_sat"] / out["post_late_orders_overall"]

    fs_del = out["post_delivered_vol_fri_sat"]
    wd_del = out["post_delivered_vol_weekday"]
    ov_del = out["post_delivered_vol_overall"]
    out["weekend_del_rel_hit_on_overall"] = (fs_del / ov_del) * out["delta_del_rel_fri_sat"]
    out["weekday_del_rel_hit_on_overall"] = (wd_del / ov_del) * out["delta_del_rel_weekday"]
    out["overall_del_rel_if_frisat_held"] = out["post_del_rel_overall"] - out["weekend_del_rel_hit_on_overall"]

    ordered = list(WIDE_ID) + [
        "pre_vol_overall",
        "post_vol_overall",
        "pre_vol_fri_sat",
        "post_vol_fri_sat",
        "pre_vol_weekday",
        "post_vol_weekday",
        "post_fri_sat_vol_share",
        "post_weekday_vol_share",
        "pre_ifr_overall",
        "post_ifr_overall",
        "delta_ifr_overall",
        "pre_ifr_fri_sat",
        "post_ifr_fri_sat",
        "delta_ifr_fri_sat",
        "pre_ifr_weekday",
        "post_ifr_weekday",
        "delta_ifr_weekday",
        "overall_ifr_if_frisat_held",
        "weekend_ifr_hit_on_overall",
        "weekday_ifr_hit_on_overall",
        "extra_late_fri_sat",
        "extra_late_weekday",
        "post_late_orders_overall",
        "post_late_orders_fri_sat",
        "post_late_orders_weekday",
        "post_late_share_fri_sat",
        "pre_del_rel_overall",
        "post_del_rel_overall",
        "delta_del_rel_overall",
        "pre_del_rel_fri_sat",
        "post_del_rel_fri_sat",
        "delta_del_rel_fri_sat",
        "pre_del_rel_weekday",
        "post_del_rel_weekday",
        "delta_del_rel_weekday",
        "overall_del_rel_if_frisat_held",
        "weekend_del_rel_hit_on_overall",
        "weekday_del_rel_hit_on_overall",
        "pre_o2d_det_overall",
        "post_o2d_det_overall",
        "delta_o2d_det_overall",
        "pre_o2d_det_fri_sat",
        "post_o2d_det_fri_sat",
        "delta_o2d_det_fri_sat",
        "pre_o2d_det_weekday",
        "post_o2d_det_weekday",
        "delta_o2d_det_weekday",
        "pre_det_del_rel_overall",
        "post_det_del_rel_overall",
        "delta_det_del_rel_overall",
        "pre_det_del_rel_fri_sat",
        "post_det_del_rel_fri_sat",
        "delta_det_del_rel_fri_sat",
        "pre_det_del_rel_weekday",
        "post_det_del_rel_weekday",
        "delta_det_del_rel_weekday",
        "pre_o2d_actual_less5_overall",
        "post_o2d_actual_less5_overall",
        "delta_o2d_actual_less5_overall",
        "pre_o2d_actual_less5_fri_sat",
        "post_o2d_actual_less5_fri_sat",
        "delta_o2d_actual_less5_fri_sat",
        "pre_o2d_actual_less5_weekday",
        "post_o2d_actual_less5_weekday",
        "delta_o2d_actual_less5_weekday",
    ]
    out = out[[c for c in ordered if c in out.columns]]
    out["weekend_ifr_hit_on_overall"] = pd.to_numeric(out["weekend_ifr_hit_on_overall"], errors="coerce")
    parts = []
    for wave in ("wave1_jun_jul", "wave2_aug"):
        parts.append(
            out[out["wave"] == wave].sort_values("weekend_ifr_hit_on_overall", ascending=True, na_position="last")
        )
    parts.append(
        out[~out["wave"].isin(["wave1_jun_jul", "wave2_aug"])].sort_values(
            "weekend_ifr_hit_on_overall", ascending=True, na_position="last"
        )
    )
    return pd.concat(parts, ignore_index=True)


def _header_fill():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    return {
        "font": Font(bold=True, color="FFFFFF", name="Calibri", size=10),
        "fill": PatternFill("solid", fgColor="1A365D"),
        "align": Alignment(wrap_text=True, vertical="center", horizontal="center"),
        "border": Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        ),
    }


def _style_sheet(ws, df: pd.DataFrame) -> None:
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    hdr = _header_fill()
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 40
    for col_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(1, col_idx)
        display = EXCEL_HEADERS.get(col, col)
        cell.value = display
        cell.font = hdr["font"]
        cell.fill = hdr["fill"]
        cell.alignment = hdr["align"]
        cell.border = hdr["border"]
        letter = get_column_letter(col_idx)
        width = min(36, max(12, len(str(display)) + 2))
        if col in {"su_name", "parent_su_name"}:
            width = 36
        ws.column_dimensions[letter].width = width

        metric = col
        for prefix in ("pre_", "post_", "delta_"):
            if metric.startswith(prefix):
                metric = metric[len(prefix) :]
                break
        if col.startswith("delta_") and metric in RATE_COLS:
            fmt = "+0.0%;-0.0%;0.0%"
        elif metric in RATE_COLS and not col.startswith("delta_"):
            fmt = "0.0%"
        elif col.startswith("delta_") and metric in DAY_COLS:
            fmt = "+0.00;-0.00;0.00"
        elif metric in DAY_COLS:
            fmt = "0.00"
        elif metric in COUNT_COLS:
            fmt = "#,##0"
        else:
            fmt = None
        if fmt:
            for row in range(2, ws.max_row + 1):
                ws.cell(row, col_idx).number_format = fmt

        if col.startswith("delta_") and metric in RATE_COLS | DAY_COLS and ws.max_row >= 2:
            # Faster stated speed is negative (green). Badge/IFR increase is positive (green).
            if metric in DAY_COLS:
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    ColorScaleRule(
                        start_type="num",
                        start_value=-0.5,
                        start_color="1E8449",
                        mid_type="num",
                        mid_value=0,
                        mid_color="FFFFFF",
                        end_type="num",
                        end_value=0.5,
                        end_color="C0392B",
                    ),
                )
            else:
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    ColorScaleRule(
                        start_type="num",
                        start_value=-0.1,
                        start_color="C0392B",
                        mid_type="num",
                        mid_value=0,
                        mid_color="FFFFFF",
                        end_type="num",
                        end_value=0.1,
                        end_color="1E8449",
                    ),
                )


def _write_summary_sheet(ws, rollup: pd.DataFrame, roster: pd.DataFrame) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    title = Font(name="Calibri", size=16, bold=True, color="1A365D")
    section = Font(name="Calibri", size=12, bold=True, color="1A365D")
    body = Font(name="Calibri", size=11)
    hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1A365D")

    ws["A1"] = "Weekend MSBD enablement — supplier pre/post workbook"
    ws["A1"].font = title
    ws.merge_cells("A1:G1")
    ws["A2"] = (
        "Generated 2026-08-31. Pre = 6 weeks before each warehouse’s first weekend-MSBD week. "
        "Post = enable week through 2026-08-16. Stated O2S = o2sumsbd (order date to supplier MSBD). "
        "Fast badge = stated O2D ≤ 5 only (customer badge). o2d_actual_less5 is actual O2D ≤ 5, not a badge. "
        "Deterministic O2D / reliability use det_delivery_date and det_del_rel. Negative day deltas = faster."
    )
    ws["A2"].font = body
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 48

    def write_block(start_row: int, heading: str, pre: pd.Series, post: pd.Series) -> int:
        ws.cell(start_row, 1, heading).font = section
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
        headers = ["Metric", "Pre", "Post", "Change", "Notes"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(start_row + 1, i, h)
            c.font = hdr_font
            c.fill = hdr_fill
        rows = [
            ("Suppliers", int(pre["n_suppliers"]), int(post["n_suppliers"]), None, None),
            ("Volume (distinct ops)", int(pre["vol"]), int(post["vol"]), None, None),
            ("Stated O2S (days)", pre["o2s_stated"], post["o2s_stated"], post["o2s_stated"] - pre["o2s_stated"], "o2sumsbd; lower = faster promise"),
            ("1-day stated O2S", pre["o2s_stated_1"], post["o2s_stated_1"], post["o2s_stated_1"] - pre["o2s_stated_1"], "o2s_stated_1 share"),
            ("Stated O2D (days)", pre["o2d_stated"], post["o2d_stated"], post["o2d_stated"] - pre["o2d_stated"], "lower = faster customer promise"),
            ("Fast badge (stated O2D ≤ 5)", pre["fast_badge"], post["fast_badge"], post["fast_badge"] - pre["fast_badge"], "o2d_stated_5 — badge is stated only"),
            ("Actual O2D ≤ 5", pre["o2d_actual_less5"], post["o2d_actual_less5"], post["o2d_actual_less5"] - pre["o2d_actual_less5"], "o2d_actual_5 — not a badge"),
            ("Deterministic O2D (days)", pre["o2d_det"], post["o2d_det"], post["o2d_det"] - pre["o2d_det"], "order date to det_delivery_date"),
            ("Actual O2S (days)", pre["o2s_actual"], post["o2s_actual"], post["o2s_actual"] - pre["o2s_actual"], None),
            ("Actual O2D (days)", pre["o2d_actual"], post["o2d_actual"], post["o2d_actual"] - pre["o2d_actual"], None),
            ("IFR", pre["ifr"], post["ifr"], post["ifr"] - pre["ifr"], None),
            ("Delivery reliability", pre["del_rel"], post["del_rel"], post["del_rel"] - pre["del_rel"], "delivered orders only"),
            ("Deterministic reliability", pre["det_del_rel"], post["det_del_rel"], post["det_del_rel"] - pre["det_del_rel"], "det_del_rel where det_delivery_date is set"),
            ("Weekend MSBD share", pre["weekend_msbd"], post["weekend_msbd"], post["weekend_msbd"] - pre["weekend_msbd"], None),
            ("Weekend ship (Sat+Sun)", pre["weekend_ship"], post["weekend_ship"], post["weekend_ship"] - pre["weekend_ship"], "induction_dow_adj 1 or 7"),
            ("Saturday induction", pre["sat_ship"], post["sat_ship"], post["sat_ship"] - pre["sat_ship"], None),
            ("Sunday induction", pre["sun_ship"], post["sun_ship"], post["sun_ship"] - pre["sun_ship"], None),
        ]
        rate_labels = {
            "1-day stated O2S",
            "Fast badge (stated O2D ≤ 5)",
            "Actual O2D ≤ 5",
            "IFR",
            "Delivery reliability",
            "Deterministic reliability",
            "Weekend MSBD share",
            "Weekend ship (Sat+Sun)",
            "Saturday induction",
            "Sunday induction",
        }
        day_labels = {
            "Stated O2S (days)",
            "Stated O2D (days)",
            "Actual O2S (days)",
            "Actual O2D (days)",
            "Deterministic O2D (days)",
        }
        for i, (label, a, b, d, note) in enumerate(rows):
            r = start_row + 2 + i
            ws.cell(r, 1, label).font = body
            ws.cell(r, 2, a)
            ws.cell(r, 3, b)
            if d is not None:
                ws.cell(r, 4, d)
            if note:
                ws.cell(r, 5, note).font = Font(name="Calibri", size=9, italic=True, color="666666")
            if label in rate_labels:
                ws.cell(r, 2).number_format = "0.0%"
                ws.cell(r, 3).number_format = "0.0%"
                if d is not None:
                    ws.cell(r, 4).number_format = "+0.0%;-0.0%;0.0%"
            elif label in day_labels:
                ws.cell(r, 2).number_format = "0.00"
                ws.cell(r, 3).number_format = "0.00"
                if d is not None:
                    ws.cell(r, 4).number_format = "+0.00;-0.00;0.00"
            elif label in {"Suppliers", "Volume (distinct ops)"}:
                ws.cell(r, 2).number_format = "#,##0"
                ws.cell(r, 3).number_format = "#,##0"
        return start_row + 2 + len(rows) + 2

    def write_slice_compare(start_row: int, heading: str, wave: str) -> int:
        fs_pre, fs_post = slice_pair(rollup, wave, "fri_sat")
        wd_pre, wd_post = slice_pair(rollup, wave, "weekday")
        all_pre, all_post = slice_pair(rollup, wave, "all")
        fs_share = fs_post["vol"] / all_post["vol"]
        wd_share = wd_post["vol"] / all_post["vol"]
        weekend_ifr_hit = fs_share * (fs_post["ifr"] - fs_pre["ifr"])
        weekday_ifr_hit = wd_share * (wd_post["ifr"] - wd_pre["ifr"])
        fs_dshare = fs_post["delivered_vol"] / all_post["delivered_vol"]
        wd_dshare = wd_post["delivered_vol"] / all_post["delivered_vol"]
        weekend_dr_hit = fs_dshare * (fs_post["del_rel"] - fs_pre["del_rel"])
        weekday_dr_hit = wd_dshare * (wd_post["del_rel"] - wd_pre["del_rel"])

        ws.cell(start_row, 1, heading).font = section
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
        headers = [
            "Slice",
            "Pre IFR",
            "Post IFR",
            "Δ IFR",
            "Pre DR",
            "Post DR",
            "Δ DR",
            "Pre det O2D",
            "Post det O2D",
            "Δ det O2D",
            "Pre det rel",
            "Post det rel",
            "Δ det rel",
        ]
        for i, h in enumerate(headers, 1):
            c = ws.cell(start_row + 1, i, h)
            c.font = hdr_font
            c.fill = hdr_fill
        rows = [
            ("Friday/Saturday placed", fs_pre, fs_post),
            ("Monday–Thursday placed", wd_pre, wd_post),
            ("Overall (all days)", all_pre, all_post),
        ]
        rate_cols = {2, 3, 4, 5, 6, 7, 11, 12, 13}
        day_cols = {8, 9, 10}
        for i, (label, pre, post) in enumerate(rows):
            rr = start_row + 2 + i
            ws.cell(rr, 1, label).font = body
            vals = [
                pre["ifr"],
                post["ifr"],
                post["ifr"] - pre["ifr"],
                pre["del_rel"],
                post["del_rel"],
                post["del_rel"] - pre["del_rel"],
                pre["o2d_det"],
                post["o2d_det"],
                post["o2d_det"] - pre["o2d_det"],
                pre["det_del_rel"],
                post["det_del_rel"],
                post["det_del_rel"] - pre["det_del_rel"],
            ]
            for c, v in enumerate(vals, 2):
                ws.cell(rr, c, v)
                if c in rate_cols:
                    ws.cell(rr, c).number_format = "+0.0%;-0.0%;0.0%" if c in {4, 7, 13} else "0.0%"
                elif c in day_cols:
                    ws.cell(rr, c).number_format = "+0.00;-0.00;0.00" if c == 10 else "0.00"

        ws.cell(
            start_row + 5,
            1,
            "Overall if Fri/Sat held at pre",
        ).font = body
        ws.cell(start_row + 5, 3, all_post["ifr"] - weekend_ifr_hit).number_format = "0.0%"
        ws.cell(start_row + 5, 6, all_post["del_rel"] - weekend_dr_hit).number_format = "0.0%"
        ws.cell(start_row + 6, 1, "Fri/Sat hit on overall (post mix × Δ)").font = body
        ws.cell(start_row + 6, 4, weekend_ifr_hit).number_format = "+0.0%;-0.0%;0.0%"
        ws.cell(start_row + 6, 7, weekend_dr_hit).number_format = "+0.0%;-0.0%;0.0%"
        ws.cell(start_row + 7, 1, "Mon–Thu hit on overall (post mix × Δ)").font = body
        ws.cell(start_row + 7, 4, weekday_ifr_hit).number_format = "+0.0%;-0.0%;0.0%"
        ws.cell(start_row + 7, 7, weekday_dr_hit).number_format = "+0.0%;-0.0%;0.0%"
        note_r = start_row + 8
        ws.cell(
            note_r,
            1,
            "Fri/Sat hit on overall = (post Fri/Sat volume share) × (post − pre Fri/Sat IFR). Same for DR using delivered volume. Mon–Thu is the weekday analogue.",
        ).font = Font(name="Calibri", size=9, italic=True, color="666666")
        ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=10)
        return note_r + 3

    r = 4
    w1_all_pre, w1_all_post = slice_pair(rollup, "wave1_jun_jul", "all")
    w1_fs_pre, w1_fs_post = slice_pair(rollup, "wave1_jun_jul", "fri_sat")
    w1_wd_pre, w1_wd_post = slice_pair(rollup, "wave1_jun_jul", "weekday")
    w2_all_pre, w2_all_post = slice_pair(rollup, "wave2_aug", "all")
    w2_fs_pre, w2_fs_post = slice_pair(rollup, "wave2_aug", "fri_sat")
    w2_wd_pre, w2_wd_post = slice_pair(rollup, "wave2_aug", "weekday")

    r = write_slice_compare(r, "Wave 1 — IFR / DR / det O2D by order day (how weekend MSBD hits overall)", "wave1_jun_jul")
    r = write_slice_compare(r, "Wave 2 — IFR / DR / det O2D by order day (early 2-week post)", "wave2_aug")
    r = write_block(
        r,
        "Wave 1 — OVERALL (all order days): stated O2S / O2D and badging",
        w1_all_pre,
        w1_all_post,
    )
    r = write_block(r, "Wave 1 — Friday/Saturday placed orders", w1_fs_pre, w1_fs_post)
    r = write_block(r, "Wave 1 — Monday–Thursday placed (same warehouses)", w1_wd_pre, w1_wd_post)
    r = write_block(
        r,
        "Wave 2 (Aug 2 sprint, early 2-week post) — OVERALL stated speed & badging",
        w2_all_pre,
        w2_all_post,
    )
    r = write_block(r, "Wave 2 — Friday/Saturday placed orders", w2_fs_pre, w2_fs_post)
    r = write_block(r, "Wave 2 — Monday–Thursday placed (same warehouses)", w2_wd_pre, w2_wd_post)

    ws.cell(r, 1, "Sheets").font = section
    notes = [
        "FriSat_vs_Weekday — warehouse-level Mon–Thu vs Fri/Sat vs overall IFR, DR, det O2D; weekend hit on overall",
        "Supplier_FriSat — every enabled warehouse, Fri/Sat orders, pre / post / delta",
        "Supplier_Overall — every enabled warehouse, all order days (speed & badge view)",
        "Supplier_Weekday — Monday–Thursday at the same warehouses",
        "Wave1_FriSat — the original 12 June/July warehouses only",
        "Roster — enable week and wave assignment",
        "o2d_actual_less5 = actual O2D ≤ 5 (not a badge). Fast badge = stated O2D ≤ 5 only.",
        "Deterministic O2D = days from order date to det_delivery_date. Det reliability = det_del_rel.",
    ]
    for i, line in enumerate(notes):
        ws.cell(r + 1 + i, 1, line).font = body
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 42
    for col in ("F", "G", "H", "I", "J", "K", "L", "M"):
        ws.column_dimensions[col].width = 13


def write_excel(
    roster: pd.DataFrame,
    rollup: pd.DataFrame,
    supplier: pd.DataFrame,
    supplier_all: pd.DataFrame,
    control: pd.DataFrame,
) -> Path:
    path = OUT / "weekend_shipping_pre_post_supplier.xlsx"
    fri = wide_supplier(supplier, "fri_sat")
    weekday = wide_supplier(supplier, "weekday")
    overall = wide_supplier(supplier_all, "all")
    impact = build_ifr_dr_impact(supplier, supplier_all)
    wave1 = fri[fri["wave"] == "wave1_jun_jul"].copy()
    wave1_impact = impact[impact["wave"] == "wave1_jun_jul"].copy()

    sheets = {
        "FriSat_vs_Weekday": impact,
        "Wave1_FriSat_vs_Weekday": wave1_impact,
        "Supplier_FriSat": fri,
        "Supplier_Overall": overall,
        "Supplier_Weekday": weekday,
        "Wave1_FriSat": wave1,
        "Roster": roster.copy(),
        "Cohort_rollup": rollup.copy(),
        "Control": control.copy(),
    }

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            export = df.copy()
            if "enable_week" in export.columns:
                export["enable_week"] = pd.to_datetime(export["enable_week"], errors="coerce").dt.date
            export.to_excel(writer, sheet_name=name, index=False)
        wb = writer.book
        ws = wb.create_sheet("Summary", 0)
        _write_summary_sheet(ws, rollup, roster)
        for name, df in sheets.items():
            _style_sheet(wb[name], df)

    return path


def fetch() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    cache = Path("/tmp/weekend_shipping_pre_post_cache_v3.pkl")
    if cache.exists():
        print(f"Loading cached extracts from {cache}")
        roster, orders, control = pd.read_pickle(cache)
        return roster, orders, control
    print("Querying enabled roster…")
    roster = query_df(with_cte("weekend_shipping_pre_post_roster.sql"))
    print("Querying enabled orders…")
    orders = prep_orders(query_df(with_cte("weekend_shipping_pre_post_orders.sql")))
    print("Querying 24hr control…")
    control = query_df(load_sql("weekend_shipping_pre_post_control.sql"))
    pd.to_pickle((roster, orders, control), cache)
    return roster, orders, control


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    CHARTS.mkdir(parents=True, exist_ok=True)

    roster, orders, control = fetch()

    rollup = (
        orders.groupby(["wave", "order_bucket", "period"], dropna=True)
        .apply(summarize)
        .reset_index()
    )
    supplier = (
        orders.groupby(
            ["wave", "supplier_id", "su_name", "parent_su_name", "sto", "srm", "enable_week", "order_bucket", "period"],
            dropna=True,
        )
        .apply(summarize)
        .reset_index()
    )
    weekly = (
        orders.groupby(["wave", "order_bucket", "weeks_from_enable", "week_start"], dropna=True)
        .apply(summarize)
        .reset_index()
    )
    all_rollup = (
        orders.groupby(["wave", "period"], dropna=True)
        .apply(summarize)
        .reset_index()
    )
    all_rollup["order_bucket"] = "all"
    rollup = pd.concat([rollup, all_rollup], ignore_index=True)

    supplier_all = (
        orders.groupby(
            ["wave", "supplier_id", "su_name", "parent_su_name", "sto", "srm", "enable_week", "period"],
            dropna=True,
        )
        .apply(summarize)
        .reset_index()
    )
    supplier_all["order_bucket"] = "all"

    roster.to_csv(OUT / "enabled_roster.csv", index=False)
    rollup.to_csv(OUT / "cohort_rollup.csv", index=False)
    supplier.to_csv(OUT / "supplier_pre_post.csv", index=False)
    supplier_all.to_csv(OUT / "supplier_pre_post_overall.csv", index=False)
    weekly.to_csv(OUT / "weekly_relative.csv", index=False)
    control.to_csv(OUT / "control_rollup.csv", index=False)
    wide_supplier(supplier, "fri_sat").to_csv(OUT / "supplier_fri_sat_wide.csv", index=False)
    wide_supplier(supplier_all, "all").to_csv(OUT / "supplier_overall_wide.csv", index=False)
    wide_supplier(supplier, "weekday").to_csv(OUT / "supplier_weekday_wide.csv", index=False)
    build_ifr_dr_impact(supplier, supplier_all).to_csv(OUT / "supplier_ifr_dr_impact.csv", index=False)

    print("Building charts…")
    chart_grouped_bars(rollup)
    chart_weekday_vs_weekend(rollup)
    chart_sat_sun(rollup)
    chart_supplier_lollipop(supplier)
    chart_weekly(orders)
    chart_wave2(rollup)

    xlsx = write_excel(roster, rollup, supplier, supplier_all, control)

    print(f"Roster: {len(roster)} enabled warehouses")
    print(roster.groupby("wave").size().to_string())
    print(f"Wrote CSVs to {OUT}")
    print(f"Wrote Excel to {xlsx}")
    print(f"Wrote charts to {CHARTS}")


if __name__ == "__main__":
    main()
